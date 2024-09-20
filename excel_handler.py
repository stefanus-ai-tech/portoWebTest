# excel_handler.py
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

EXCEL_FILE = 'contacts.xlsx'

def save_contact_form(name, email, message):
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(['Timestamp', 'Name', 'Email', 'Message'])
        wb.save(EXCEL_FILE)
    
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.append([timestamp, name, email, message])
        wb.save(EXCEL_FILE)
        return True
    except Exception as e:
        print(f"Error saving contact form: {e}")
        return False
